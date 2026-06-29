from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

FONT = "Arial"
INK = "1E1B4B"      # deep indigo (game UI)
ACCENT = "4338CA"
WHITE = "FFFFFF"
P1 = "C6EFCE"; P2 = "FFEB9C"; P3 = "E7E6E6"

thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

# ───────────────────────── Sheet 1: Soundtrack ─────────────────────────
ws = wb.active
ws.title = "Soundtrack"

headers = ["Cue ID", "Track Name", "Game Context", "Type", "Mood / Energy",
           "Suno Style Prompt (paste-ready)", "Vocals / Lyric Seed",
           "Length & Loop", "Priority", "Notes"]

rows = [
 ["MTH-01","Title Theme — \"Return to the Meld\"","Title screen / main menu","Loop","Haunting, melancholic, low",
  "dark ambient cinematic game score, glassy detuned synth pads, deep sub bass, distant wordless ethereal female choir, 64 bpm, reverberant, slow, hopeful-sad, A minor",
  "Wordless choir. Optional refrain: \"all that is held / returns to the meld\"","~2:00, seamless loop","P1",
  "Establishes the MELD MOTIF — a slow descending 4-note phrase reused in the Mirror & Ending. This is the album's signature."],

 ["MTH-02","The Unmerged","\"YOU HAVE BEEN UNMERGED\" — class select","Loop","Cold, dissociative, searching",
  "cold ambient drone, granular glitch textures, glassy bell tones, sparse sub bass, no drums, 60 bpm, dissociative, clinical, A minor",
  "[Instrumental]","~1:30 loop","P2",
  "Identity-loss feel — unsettling but not hostile. The moment he's been pulled apart."],

 ["MTH-03","Forge Your Deck","Loadout screen before a fight","Loop","Focused, anticipatory",
  "minimal electronic underscore, soft pulsing synth arpeggio, warm pad, light ticking percussion, 90 bpm, focused, anticipatory, low tension, A minor",
  "[Instrumental]","~1:30 loop","P2",
  "Keep it understated so it doesn't fatigue on long deck-building. Slight forward pull toward the fight."],

 ["MTH-04","Battle I — The Proving","Run 1 combat (Whelp / Brute / CORE tier)","Loop","Driving, mid intensity",
  "tense electronic battle theme, pulsing arpeggiated synth, driving muted low percussion, staccato string stabs, 100 bpm, dark, propulsive, loopable combat, A minor",
  "[Instrumental]","~1:30-2:00 loop","P1",
  "Core combat bed. Steady groove for long turns. Same key as Battle II/III so they feel like one escalating piece."],

 ["MTH-05","Battle II — The Deepening","Run 2 combat","Loop","Darker, heavier",
  "heavier dark electronic combat, distorted bass synth, hybrid tribal percussion, ominous low brass swells, 110 bpm, menacing, driving, loopable, A minor",
  "[Instrumental]","~1:30-2:00 loop","P1",
  "Reuse Battle I's motif, add weight + low end. The walls are sharper now."],

 ["MTH-06","Battle III — The Reckoning","Run 3 combat (apex enemies)","Loop","Intense, urgent, full",
  "intense cinematic electronic battle, fast driving percussion, aggressive synth bass, frantic arpeggios, dissonant strings, 124 bpm, high-stakes, relentless, loopable, A minor",
  "[Instrumental]","~1:30-2:00 loop","P1",
  "Peak combat energy right before the finale. Crank tempo + dissonance."],

 ["MTH-07","The Mirror — \"Your Echo\"","10th opponent / finale (YOUR ECHO)","Loop / build","Climactic, personal, tragic",
  "climactic emotional battle score, full MELD MOTIF on strings and choir, pounding hybrid percussion, soaring detuned synth lead, 120 bpm, tragic, intimate, epic, A minor",
  "Choir. Optional: \"you wore my face / you forgot my name\"","~2:30, build then loop","P1",
  "The emotional peak — fighting yourself. Fully orchestrate the title's MELD MOTIF here."],

 ["MTH-08","Meld to All Held — Ending","Full game complete / ending screen","One-shot","Resolving, bittersweet, cathartic",
  "bittersweet orchestral-electronic resolution, MELD MOTIF resolving, warm choir, gentle piano, soft sub bass, 70 bpm, cathartic, hopeful, A minor lifting to A major",
  "Wordless choir; resolve the motif","~0:40 one-shot (no loop)","P1",
  "Pays off the title motif and the loop/memory arc. The major lift is the catharsis."],

 ["MTH-09","Form Lost — Defeat","Defeat screen (\"FORM LOST\")","One-shot","Somber, dissolving",
  "somber ambient sting, descending detuned piano, fading granular textures, hollow sub drop, no drums, 60 bpm, grief, dissolution, A minor",
  "[Instrumental]","~0:15-0:20 one-shot","P1",
  "Dissolution = the form un-melding. Short, no loop — let it decay to silence."],

 ["MTH-10","Between Runs — Evolution","Evolution + modifier / boon screens","Loop","Reflective, weighty",
  "reflective ambient interlude, sustained warm pads, distant choir, slow bell arpeggio, no percussion, 66 bpm, contemplative, weighty, A minor",
  "[Instrumental]","~1:15 loop","P2",
  "A breath between gauntlets — the \"who do you become\" choice (Deepen vs Absorb)."],

 ["MTH-11","Memory Cutscene — \"Do you know me?\"","Opponent story beats / cutscenes (letterbox)","Loop / underscore","Sparse, aching, uncanny",
  "sparse emotional underscore, lone detuned music box, soft sub drone, occasional reversed piano, no drums, 58 bpm, aching, nostalgic, uncanny, A minor",
  "[Instrumental]","~1:00 loop, very low","P2",
  "The MEMORY TIMBRE (music box). Sits UNDER dialogue — keep dynamics low. This is the heart of the opponents-trying-to-make-you-remember thread."],

 ["MTH-12","Run Cleared — Stinger","After clearing a mark / run","One-shot","Brief, earned",
  "short triumphant electronic stinger, rising synth swell, single choir hit, soft percussion flourish, resolved, A minor to C major",
  "[Instrumental]","~0:06 one-shot","P2",
  "Quick reward feedback between encounters. Generate long, trim the best 6 seconds."],

 ["MTH-13","Herald / Dialogue Bed","Herald intro & non-cutscene dialogue","Loop","Subtle, almost silent",
  "very subtle ambient bed, low sustained pad, faint shimmer, no melody, no drums, 60 bpm, neutral, spacious, A minor",
  "[Instrumental]","~1:00 loop","P3",
  "Should nearly disappear behind voice/text. Texture, not a tune."],

 ["MTH-14","Collection / Shop","Collection & fragment shop screens","Loop","Calm, curious",
  "calm ambient electronic, gentle plucked synth, warm pad, soft vinyl crackle, light shaker, 84 bpm, cozy, curious, A minor",
  "[Instrumental]","~1:15 loop","P3",
  "Relaxed browsing music — the one warm, safe room in the game."],

 ["MTH-15","Meld Success — Stinger","On a successful card meld (core mechanic)","One-shot","Bright, satisfying",
  "very short bright magical chime, ascending glassy bell triad, soft shimmer tail, satisfying, resolved",
  "[Instrumental]","~0:02 one-shot","P3",
  "Musical reward for the signature mechanic. NOTE: could also be done as a synth SFX — Suno is overkill but can work if you want it tonal."],

 ["MTH-16","The Meld — Ambient Drone","Reusable background bed under any menu","Loop","Neutral, vast, liminal",
  "deep evolving ambient drone, slow shifting glassy harmonics, sub bass, faint choir wash, no rhythm, 60 bpm, vast, liminal, A minor",
  "[Instrumental]","~2:00 seamless loop","P3",
  "Swiss-army bed. Layer under any quiet screen to unify the soundscape if a dedicated cue isn't ready."],

 ["MTH-17","The Mirror (Remembered) — NG+","Finale when cycles > 0 (he remembers)","Loop / build","Same as Mirror but knowing, resolved",
  "same as The Mirror but warmer and resolving, MELD MOTIF in major, less dissonance, choir more present, 120 bpm, tragic-hopeful, knowing, A minor to major",
  "Choir: \"I remember your name\"","~2:30","P3",
  "OPTIONAL: the finale shifts once you've looped. Ties to the cycles/memory system — pairs with MTH-08."],

 ["MTH-18","Archetype Motifs — optional layers","Short leitmotifs to flavor battle by enemy archetype","One-shot x3","Distinct per wall",
  "three short 4-bar motifs, same key A minor: PURIFIER = cold sterile glass bells; BULWARK = heavy metallic clangs; RENEWER = swelling organic breathing pads; loopable stems",
  "[Instrumental]","~0:08 each, loopable","P3",
  "ADVANCED: layer the matching motif over the run battle theme when facing that archetype. Skip unless you want deep polish."],
]

# header
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = Font(name=FONT, bold=True, color=WHITE, size=11)
    cell.fill = PatternFill("solid", fgColor=INK)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = border

prio_fill = {"P1": P1, "P2": P2, "P3": P3}
for r, row in enumerate(rows, 2):
    for c, val in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = Font(name=FONT, size=10)
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell.border = border
    pcell = ws.cell(row=r, column=9)
    pcell.fill = PatternFill("solid", fgColor=prio_fill[row[8]])
    pcell.font = Font(name=FONT, size=10, bold=True)
    pcell.alignment = Alignment(horizontal="center", vertical="center")

widths = [9, 30, 30, 13, 22, 50, 30, 17, 9, 42]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[chr(64+i)].width = w

ws.freeze_panes = "A2"
ws.row_dimensions[1].height = 30

# ───────────────────────── Sheet 2: Palette & Suno Tips ─────────────────────────
ws2 = wb.create_sheet("Palette & Suno Tips")
ws2.column_dimensions["A"].width = 26
ws2.column_dimensions["B"].width = 92

def section(ws, r, title):
    c = ws.cell(row=r, column=1, value=title)
    c.font = Font(name=FONT, bold=True, color=WHITE, size=11)
    c.fill = PatternFill("solid", fgColor=ACCENT)
    ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor=ACCENT)

def kv(ws, r, k, v):
    a = ws.cell(row=r, column=1, value=k)
    a.font = Font(name=FONT, bold=True, size=10)
    a.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    b = ws.cell(row=r, column=2, value=v)
    b.font = Font(name=FONT, size=10)
    b.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

title = ws2.cell(row=1, column=1, value="MELD TO ALL HELD — Sound Direction")
title.font = Font(name=FONT, bold=True, size=14, color=INK)

r = 3
section(ws2, r, "Overall aesthetic"); r += 1
kv(ws2, r, "Genre", "Dark ambient electronic + cinematic. Glassy detuned synths, sub bass, distant choir, low strings. Cold and reverberant, with a melancholic, looping ache — a game about memory, melding, and returning."); r += 1
kv(ws2, r, "Cohesion key", "Keep everything in A minor (or its relatives). Sharing a key lets tracks crossfade and the recurring motif stay recognizable."); r += 1
kv(ws2, r, "The MELD MOTIF", "Write ONE simple melodic phrase (a slow descending 3-4 note figure). Plant it wordless in the Title (MTH-01), orchestrate it at the Mirror (MTH-07), and resolve it at the Ending (MTH-08). That single thread makes the score feel composed, not assembled."); r += 1
kv(ws2, r, "The MEMORY timbre", "A lone detuned music box / celesta = 'he's almost remembering.' Reserve it for the cutscene beats (MTH-11) so it always signals memory."); r += 2

section(ws2, r, "Using Suno"); r += 1
kv(ws2, r, "Instrumental", "Toggle Instrumental ON (or put [Instrumental] in lyrics) for every cue except Title / Mirror / Ending, where wordless choir or a short refrain is fine."); r += 1
kv(ws2, r, "Style prompt", "Paste the 'Suno Style Prompt' cell straight into the Style box. They're already comma-keyworded the way Suno likes. Trim if it complains about length."); r += 1
kv(ws2, r, "Loops", "Suno outputs full ~2-min songs. For loop cues, find a clean 30-90s section with a steady bar count and trim/crossfade it in an editor (Audacity). Generate a few, pick the most seamless."); r += 1
kv(ws2, r, "Stingers", "For one-shots (MTH-08, 09, 12, 15) generate the full song, then cut the single swell/hit you want. Easier than asking Suno for 6 seconds."); r += 1
kv(ws2, r, "Consistency", "Reuse the same core style words across the three Battle cues (and Title/Mirror/Ending) so they sound like one record. A consistent 'persona'/seed in Suno helps too."); r += 1
kv(ws2, r, "Format for the game", "Export to a web-friendly format (.ogg or .mp3, ~128-192kbps). The game has no music system yet — these are assets for when one's added."); r += 2

section(ws2, r, "Priority"); r += 1
kv(ws2, r, "P1 — ship first", "Title, the 3 Battle themes, Mirror, Ending, Defeat. With just these the whole game has a voice."); r += 1
kv(ws2, r, "P2 — round it out", "Unmerged, Forge, Evolution, Memory cutscene, Run-clear stinger."); r += 1
kv(ws2, r, "P3 — polish / optional", "Dialogue bed, Collection, Meld chime, ambient drone, NG+ Mirror, archetype motifs."); r += 2

section(ws2, r, "Not on this list (do NOT use Suno)"); r += 1
kv(ws2, r, "Gameplay SFX", "Card play, hits, shield, poison tick, button clicks already live in src/sfx.ts as synthesized effects. Those are foley/synth territory, not Suno. This sheet is music + musical stingers only."); r += 1

for row in range(1, r+1):
    ws2.row_dimensions[row].height = 30

wb.save("/Users/omarilocal/meld/audio/MELD_suno_soundtrack.xlsx")
print("saved", len(rows), "cues")
